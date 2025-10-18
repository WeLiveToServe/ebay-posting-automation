"""
Batch-generate an eBay upload workbook from multiple agent outputs.

For each folder in batch-image-sets (or a user-specified subset), this script:
  - Reads the <folder>.txt agent output with `price ||| html ||| condition`
  - Reads the accompanying uploaded_urls.txt manifest
  - Extracts author and title metadata from the HTML
  - Builds a row using fixed constants for the minimal eBay columns

All rows are appended to a single workbook named `ebay-upl-<MM>-<DD>-<HH>-<MM>.xlsx`
written to the requested output directory (default: current directory).
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook

RESULTS_DIR = Path("batch-JSON-results")
IMAGE_ROOT = Path("batch-image-sets")
URL_MANIFEST = "uploaded_urls.txt"

HEADERS: List[str] = [
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)",
    "Category ID",
    "Category name",
    "Title",
    "Start price",
    "Quantity",
    "Item photo URL",
    "Condition ID",
    "Description",
    "Format",
    "Duration",
    "Location",
    "Shipping profile name",
    "Return profile name",
    "Payment profile name",
    "C:Author",
    "C:Book Title",
    "C:Language",
]

CONSTANTS: Dict[str, str] = {
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)": "Add",
    "Category ID": "261186",
    "Category name": "/Books & Magazines/Books",
    "Quantity": "1",
    "Format": "FixedPrice",
    "Duration": "GTC",
    "Location": "Newfields, NH",
    "Shipping profile name": "USPS Media Mail",
    "Return profile name": "Returns allowed within 30 days",
    "Payment profile name": "Immediate payment managed via eBay",
    "C:Language": "English",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build an eBay upload workbook from agent outputs and URL manifests."
    )
    parser.add_argument(
        "--folders",
        nargs="*",
        help="Specific folder names to process (default: all directories under batch-image-sets).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("."),
        help="Output directory for the workbook (default: current directory).",
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help="Append to the latest existing ebay-upl-*.xlsx in the output directory instead of creating a new file.",
    )
    return parser.parse_args()


def iter_target_folders(selected: Iterable[str] | None) -> List[str]:
    if selected:
        return sorted(selected)

    if not IMAGE_ROOT.exists():
        raise FileNotFoundError(f"Image root not found: {IMAGE_ROOT}")

    return sorted(str(path.name) for path in IMAGE_ROOT.iterdir() if path.is_dir())


def load_agent_output(folder: str) -> Tuple[str, str, str]:
    txt_path = RESULTS_DIR / f"{folder}.txt"
    if not txt_path.exists():
        raise FileNotFoundError(f"Agent output not found: {txt_path}")
    raw = txt_path.read_text(encoding="utf-8").strip()
    parts = [segment.strip() for segment in raw.split(" ||| ")]
    if len(parts) != 3:
        raise ValueError(f"Unexpected output format in {txt_path.name}")
    price, html, condition = parts
    return price, html, condition


def load_image_urls(folder: str) -> str:
    manifest_path = IMAGE_ROOT / folder / URL_MANIFEST
    if not manifest_path.exists():
        raise FileNotFoundError(f"URL manifest not found: {manifest_path}")
    content = manifest_path.read_text(encoding="utf-8").strip()
    if not content:
        raise ValueError(f"URL manifest is empty: {manifest_path}")
    return content


def extract_metadata(description_html: str) -> Tuple[str, str]:
    def extract(field: str) -> str:
        pattern = rf"<li>\s*{re.escape(field)}:\s*(.*?)</li>"
        match = re.search(pattern, description_html, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            return ""
        value = re.sub(r"<.*?>", "", match.group(1)).strip()
        return value

    author = extract("Author")
    title = extract("Title")
    return author, title


def truncate(text: str, limit: int = 60) -> str:
    if len(text) <= limit:
        return text
    if limit <= 3:
        return text[:limit]
    return text[: limit - 3].rstrip() + "..."


def build_row(price: str, condition: str, description_html: str, image_urls: str) -> Dict[str, str]:
    author, book_title = extract_metadata(description_html)
    clean_title = truncate(book_title or "Untitled Listing")

    row = {header: "" for header in HEADERS}
    row.update(CONSTANTS)
    row["Start price"] = price
    row["Condition ID"] = condition
    row["Description"] = description_html
    row["Item photo URL"] = image_urls
    row["C:Author"] = author or ""
    row["C:Book Title"] = book_title or ""
    row["Title"] = clean_title
    return row


def prepare_workbook(output_dir: Path, append: bool) -> Tuple[Workbook, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    if append:
        candidates = sorted(output_dir.glob("ebay-upl-*.xlsx"))
        if not candidates:
            raise FileNotFoundError("Append requested but no existing ebay-upl-*.xlsx found in output directory.")
        latest_path = candidates[-1]
        wb = load_workbook(latest_path)
        ws = wb.active
        if [cell.value for cell in ws[1]] != HEADERS:
            raise ValueError(f"Existing workbook {latest_path} does not match expected headers.")
        return wb, latest_path

    timestamp = datetime.now()
    filename = f"ebay-upl-{timestamp:%m-%d-%H-%M}.xlsx"
    workbook_path = output_dir / filename
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    return wb, workbook_path


def append_rows(
    wb: Workbook,
    workbook_path: Path,
    rows: List[Dict[str, str]],
) -> None:
    ws = wb.active
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    wb.save(workbook_path)


def main() -> None:
    args = parse_args()
    folders = iter_target_folders(args.folders)

    if not folders:
        print("No folders to process.")
        return

    workbook, workbook_path = prepare_workbook(args.output, args.append)
    rows: List[Dict[str, str]] = []
    processed: List[str] = []

    for folder in folders:
        try:
            price, html, condition = load_agent_output(folder)
            image_urls = load_image_urls(folder)
            row = build_row(price, condition, html, image_urls)
            rows.append(row)
            processed.append(folder)
        except Exception as exc:
            print(f"Skipping {folder}: {exc}")

    if not rows:
        print("No valid rows generated.")
        return

    append_rows(workbook, workbook_path, rows)
    print(f"Wrote {len(rows)} row(s) to {workbook_path.name}")
    if processed:
        print("Processed folders:")
        for folder in processed:
            print(f" - {folder}")


if __name__ == "__main__":
    main()

