"""
Append queued agent JSON outputs into the eBay listings workbook.

This script looks inside a specified directory (via --input) for `.json` files,
converts each into a row, appends the rows to `ebay-auto-listings.xlsx`, and then
moves processed JSON files into a `processed` subdirectory within the input folder.
"""

from __future__ import annotations

import argparse
import json
import shutil
import warnings
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Default directory paths
DEFAULT_QUEUE_DIR = Path("queue-JSONs-to-excel")
WORKBOOK_PATH = Path("ebay-auto-listings.xlsx")

REQUIRED_HEADERS = [
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)",
    "Custom label (SKU)",
    "Category ID",
    "Category name",
    "Title",
    "Relationship",
    "Relationship details",
    "Schedule Time",
    "P:ISBN",
    "P:EPID",
    "Start price",
    "Quantity",
    "Item photo URL",
    "VideoID",
    "Condition ID",
    "Description",
    "Format",
    "Duration",
    "Buy It Now price",
    "Best Offer Enabled",
    "Best Offer Auto Accept Price",
    "Minimum Best Offer Price",
    "Immediate pay required",
    "Location",
    "Shipping service 1 option",
    "Shipping service 1 cost",
    "Shipping service 1 priority",
    "Shipping service 2 option",
    "Shipping service 2 cost",
    "Shipping service 2 priority",
    "Max dispatch time",
    "Returns accepted option",
    "Returns within option",
    "Refund option",
    "Return shipping cost paid by",
    "Shipping profile name",
    "Return profile name",
    "Payment profile name",
    "C:Author",
    "C:Book Title",
    "C:Language",
    "C:Topic",
    "C:Publisher",
    "C:Format",
    "C:Genre",
    "C:Book Series",
    "C:Publication Year",
    "C:Original Language",
    "C:Features",
    "C:Type",
    "C:Country/Region of Manufacture",
    "C:Edition",
    "C:Narrative Type",
    "C:Signed",
    "C:Intended Audience",
    "C:Binding",
    "C:Subject",
    "C:Special Attributes",
    "Product Safety Pictograms",
    "Product Safety Statements",
    "Product Safety Component",
    "Regulatory Document Ids",
    "Manufacturer Name",
    "Manufacturer AddressLine1",
    "Manufacturer AddressLine2",
    "Manufacturer City",
    "Manufacturer Country",
    "Manufacturer PostalCode",
    "Manufacturer StateOrProvince",
    "Manufacturer Phone",
    "Manufacturer Email",
    "Manufacturer ContactURL",
    "Responsible Person 1",
    "Responsible Person 1 Type",
    "Responsible Person 1 AddressLine1",
    "Responsible Person 1 AddressLine2",
    "Responsible Person 1 City",
    "Responsible Person 1 Country",
    "Responsible Person 1 PostalCode",
    "Responsible Person 1 StateOrProvince",
    "Responsible Person 1 Phone",
    "Responsible Person 1 Email",
    "Responsible Person 1 ContactURL",
]

DEFAULT_VALUES = {
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)": "Add",
    "Category ID": "261186",
    "Category name": "/Books & Magazines/Books",
    "Start price": "5.00",
    "Quantity": 1,
    "Item photo URL": "https://keith-ebay-images.s3.us-east-2.amazonaws.com/IMG_4929.JPG",
    "Condition ID": "5000-Good",
    "Format": "FixedPrice",
    "Duration": "GTC",
    "Location": "Newfields, NH",
    "Shipping profile name": "USPS Media Mail",
    "Return profile name": "Returns allowed within 30 days",
    "Payment profile name": "Immediate payment managed via eBay",
    "C:Language": "English",
}

FORCED_BLANK_HEADERS = {
    "Max dispatch time",
    "Returns accepted option",
    "Returns within option",
    "Refund option",
    "Return shipping cost paid by",
}

JSON_FIELD_MAP = {
    "title": "Title",
    "author": "C:Author",
    "edition": "C:Edition",
    "year": "C:Publication Year",
    "publisher": "C:Publisher",
}


# Parse command-line arguments - input directory path is specified here via --input flag
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append queued JSON files to eBay listings workbook with configurable input directory"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_QUEUE_DIR,
        help=f"Directory containing JSON files to process (default: {DEFAULT_QUEUE_DIR})"
    )
    return parser.parse_args()


# Collect all JSON files from the specified queue directory, sorted by modification time
def collect_queue(queue_dir: Path) -> list[Path]:
    if not queue_dir.exists():
        return []
    return sorted(queue_dir.glob("*.json"), key=lambda p: p.stat().st_mtime)


# Load and parse a single JSON file
def load_json(path: Path) -> Mapping[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, Mapping):
        raise ValueError("JSON payload must be an object")
    return data


# Normalize text by replacing common encoding artifacts and stripping whitespace
def normalise_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\uFFFD": "'",
        "’": "'",
        "“": '"',
        "”": '"',
        "–": "-",
        "—": "-",
        "…": "...",
        "•": "-",
        "©": "(c)",
        "�": "'",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text.strip()


# Build a formatted description from the JSON payload's blurb, condition, and details fields
def build_description(payload: Mapping[str, Any]) -> str:
    blurb = normalise_text(payload.get("blurb"))
    condition = normalise_text(payload.get("condition"))
    details = normalise_text(payload.get("details"))
    sections: list[str] = []
    if blurb:
        sections.append(blurb)
    if condition:
        sections.append(f"Condition Notes:\n{condition}")
    if details:
        sections.append(f"Collector Details:\n{details}")
    return "\n\n".join(sections)


# Read the header row from the Excel worksheet
def read_headers(sheet: Worksheet) -> list[str]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    return [cell.value if isinstance(cell.value, str) else "" for cell in header_row]


# Find the next available row for insertion by checking the Title column
def find_insert_row(sheet: Worksheet, title_col: int) -> int:
    row_index = sheet.max_row + 1
    while sheet.cell(row=row_index, column=title_col).value not in (None, ""):
        row_index += 1
    return row_index


# Build a row dictionary from the JSON payload, applying defaults and field mappings
def build_row(header_order: list[str], payload: Mapping[str, Any]) -> dict[str, Any]:
    row = {header: "" for header in header_order}
    row.update({key: value for key, value in DEFAULT_VALUES.items() if key in row})

    payload_lower = {str(key).lower(): value for key, value in payload.items()}
    for json_field, header in JSON_FIELD_MAP.items():
        if header in row:
            row[header] = normalise_text(payload_lower.get(json_field))

    if "Title" in row:
        row["Title"] = row.get("Title", "")
    if "C:Book Title" in row:
        row["C:Book Title"] = row.get("Title", "")
    if "Description" in row:
        row["Description"] = build_description(payload_lower)

    for blank_header in FORCED_BLANK_HEADERS:
        if blank_header in row:
            row[blank_header] = ""

    return row


# Append a row of data to the worksheet at the next available position
def append_row(sheet: Worksheet, headers: list[str], row_values: Mapping[str, Any]) -> int:
    if "Title" not in headers:
        raise ValueError("Workbook header is missing 'Title'.")
    title_col_index = headers.index("Title") + 1
    target_row = find_insert_row(sheet, title_col_index)

    for col_index, header in enumerate(headers, start=1):
        value = row_values.get(header, "")
        cell = sheet.cell(row=target_row, column=col_index)
        if value in ("", None):
            cell.value = None
            continue
        if header == "Start price":
            try:
                cell.value = float(value)
                continue
            except (TypeError, ValueError):
                pass
        if header == "Quantity":
            try:
                cell.value = int(value)
                continue
            except (TypeError, ValueError):
                pass
        cell.value = value

    return target_row


# Main processing function: collect JSONs, append to workbook, move processed files
def process_queue(queue_dir: Path) -> None:
    json_files = collect_queue(queue_dir)
    if not json_files:
        print(f"No JSON files found in {queue_dir}.")
        return

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if "Listings" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a 'Listings' sheet.")
    sheet = wb["Listings"]
    headers = read_headers(sheet)

    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Workbook is missing expected columns: {missing[:5]}...")

    processed: list[Path] = []
    for json_path in json_files:
        try:
            payload = load_json(json_path)
        except Exception as exc:
            print(f"Skipping {json_path.name}: {exc}")
            continue

        row_data = build_row(headers, payload)
        append_row(sheet, headers, row_data)
        processed.append(json_path)
        print(f"Appended {json_path.name}")

    if processed:
        wb.save(WORKBOOK_PATH)
        # Create processed subdirectory within the input directory
        processed_dir = queue_dir / "processed"
        processed_dir.mkdir(parents=True, exist_ok=True)
        for path in processed:
            destination = processed_dir / path.name
            try:
                shutil.move(str(path), destination)
            except Exception as exc:
                print(f"Warning: failed to move {path.name}: {exc}")
        print(f"Appended {len(processed)} listing(s) to {WORKBOOK_PATH}")
    else:
        print("No listings appended.")


# Main entry point: parse arguments and begin processing
def main() -> None:
    args = parse_args()
    
    # Convert the input argument to a Path object and validate it exists
    queue_dir = args.input
    if not queue_dir.exists():
        print(f"Error: Input directory not found: {queue_dir}")
        return
    if not queue_dir.is_dir():
        print(f"Error: Input path is not a directory: {queue_dir}")
        return
    
    # Process all JSON files in the specified directory
    process_queue(queue_dir)


if __name__ == "__main__":
    main()