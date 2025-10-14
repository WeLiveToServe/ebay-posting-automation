"""
Append queued agent JSON outputs into the eBay listings workbook.

This script looks inside `queue-JSONs-to-excel` for `.json` files, converts
each into a row, appends the rows to `ebay-auto-listings.xlsx`, and then moves
processed JSON files into `queue-JSONs-to-excel/processed`.
"""

from __future__ import annotations

import json
import shutil
import warnings
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

QUEUE_DIR = Path("queue-JSONs-to-excel")
PROCESSED_DIR = QUEUE_DIR / "processed"
WORKBOOK_PATH = Path("ebay-auto-listings.xlsx")

REQUIRED_HEADERS = [
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)",
    "Custom label (SKU)",
    "Category ID",
    "Category name",
    "Title",
    "Relationship",
    "Relationship details",
    "Schedule Time",
    "P:ISBN",
    "P:EPID",
    "Start price",
    "Quantity",
    "Item photo URL",
    "VideoID",
    "Condition ID",
    "Description",
    "Format",
    "Duration",
    "Buy It Now price",
    "Best Offer Enabled",
    "Best Offer Auto Accept Price",
    "Minimum Best Offer Price",
    "Immediate pay required",
    "Location",
    "Shipping service 1 option",
    "Shipping service 1 cost",
    "Shipping service 1 priority",
    "Shipping service 2 option",
    "Shipping service 2 cost",
    "Shipping service 2 priority",
    "Max dispatch time",
    "Returns accepted option",
    "Returns within option",
    "Refund option",
    "Return shipping cost paid by",
    "Shipping profile name",
    "Return profile name",
    "Payment profile name",
    "C:Author",
    "C:Book Title",
    "C:Language",
    "C:Topic",
    "C:Publisher",
    "C:Format",
    "C:Genre",
    "C:Book Series",
    "C:Publication Year",
    "C:Original Language",
    "C:Features",
    "C:Type",
    "C:Country/Region of Manufacture",
    "C:Edition",
    "C:Narrative Type",
    "C:Signed",
    "C:Intended Audience",
    "C:Binding",
    "C:Subject",
    "C:Special Attributes",
    "Product Safety Pictograms",
    "Product Safety Statements",
    "Product Safety Component",
    "Regulatory Document Ids",
    "Manufacturer Name",
    "Manufacturer AddressLine1",
    "Manufacturer AddressLine2",
    "Manufacturer City",
    "Manufacturer Country",
    "Manufacturer PostalCode",
    "Manufacturer StateOrProvince",
    "Manufacturer Phone",
    "Manufacturer Email",
    "Manufacturer ContactURL",
    "Responsible Person 1",
    "Responsible Person 1 Type",
    "Responsible Person 1 AddressLine1",
    "Responsible Person 1 AddressLine2",
    "Responsible Person 1 City",
    "Responsible Person 1 Country",
    "Responsible Person 1 PostalCode",
    "Responsible Person 1 StateOrProvince",
    "Responsible Person 1 Phone",
    "Responsible Person 1 Email",
    "Responsible Person 1 ContactURL",
]

DEFAULT_VALUES = {
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)": "Add",
    "Category ID": "261186",
    "Category name": "/Books & Magazines/Books",
    "Start price": "5.00",
    "Quantity": 1,
    "Item photo URL": "https://keith-ebay-images.s3.us-east-2.amazonaws.com/IMG_4929.JPG",
    "Condition ID": "5000-Good",
    "Format": "FixedPrice",
    "Duration": "GTC",
    "Location": "Newfields, NH",
    "Shipping profile name": "USPS Media Mail",
    "Return profile name": "Returns allowed within 30 days",
    "Payment profile name": "Immediate payment managed via eBay",
    "C:Language": "English",
}

FORCED_BLANK_HEADERS = {
    "Max dispatch time",
    "Returns accepted option",
    "Returns within option",
    "Refund option",
    "Return shipping cost paid by",
}

JSON_FIELD_MAP = {
    "title": "Title",
    "author": "C:Author",
    "edition": "C:Edition",
    "year": "C:Publication Year",
    "publisher": "C:Publisher",
}


def collect_queue() -> list[Path]:
    if not QUEUE_DIR.exists():
        return []
    return sorted(QUEUE_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime)


def load_json(path: Path) -> Mapping[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, Mapping):
        raise ValueError("JSON payload must be an object")
    return data


def normalise_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\uFFFD": "'",
        "’": "'",
        "“": '"',
        "”": '"',
        "–": "-",
        "—": "-",
        "…": "...",
        "•": "-",
        "©": "(c)",
        "�": "'",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text.strip()


def build_description(payload: Mapping[str, Any]) -> str:
    blurb = normalise_text(payload.get("blurb"))
    condition = normalise_text(payload.get("condition"))
    details = normalise_text(payload.get("details"))
    sections: list[str] = []
    if blurb:
        sections.append(blurb)
    if condition:
        sections.append(f"Condition Notes:\n{condition}")
    if details:
        sections.append(f"Collector Details:\n{details}")
    return "\n\n".join(sections)


def read_headers(sheet: Worksheet) -> list[str]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    return [cell.value if isinstance(cell.value, str) else "" for cell in header_row]


def find_insert_row(sheet: Worksheet, title_col: int) -> int:
    row_index = sheet.max_row + 1
    while sheet.cell(row=row_index, column=title_col).value not in (None, ""):
        row_index += 1
    return row_index


def truncate_for_excel(value: Any, limit: int) -> str:
    text = normalise_text(value)
    if len(text) <= limit:
        return text
    if limit <= 3:
        return text[:limit]
    return text[: limit - 3].rstrip() + "..."


def build_row(header_order: list[str], payload: Mapping[str, Any]) -> dict[str, Any]:
    row = {header: "" for header in header_order}
    row.update({key: value for key, value in DEFAULT_VALUES.items() if key in row})

    payload_lower = {str(key).lower(): value for key, value in payload.items()}
    for json_field, header in JSON_FIELD_MAP.items():
        if header in row:
            row[header] = normalise_text(payload_lower.get(json_field))

    if "Title" in row:
        row["Title"] = row.get("Title", "")
    if "C:Book Title" in row:
        row["C:Book Title"] = truncate_for_excel(row.get("Title", ""), 50)
    if "C:Author" in row:
        source_author = row.get("C:Author") or payload_lower.get("author")
        row["C:Author"] = truncate_for_excel(source_author, 50)
    if "Description" in row:
        row["Description"] = build_description(payload_lower)

    for blank_header in FORCED_BLANK_HEADERS:
        if blank_header in row:
            row[blank_header] = ""

    return row


def append_row(sheet: Worksheet, headers: list[str], row_values: Mapping[str, Any]) -> int:
    if "Title" not in headers:
        raise ValueError("Workbook header is missing 'Title'.")
    title_col_index = headers.index("Title") + 1
    target_row = find_insert_row(sheet, title_col_index)

    for col_index, header in enumerate(headers, start=1):
        value = row_values.get(header, "")
        cell = sheet.cell(row=target_row, column=col_index)
        if value in ("", None):
            cell.value = None
            continue
        if header == "Start price":
            try:
                cell.value = float(value)
                continue
            except (TypeError, ValueError):
                pass
        if header == "Quantity":
            try:
                cell.value = int(value)
                continue
            except (TypeError, ValueError):
                pass
        cell.value = value

    return target_row


def process_queue() -> None:
    json_files = collect_queue()
    if not json_files:
        print("No JSON files found in queue-JSONs-to-excel.")
        return

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if "Listings" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a 'Listings' sheet.")
    sheet = wb["Listings"]
    headers = read_headers(sheet)

    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Workbook is missing expected columns: {missing[:5]}...")

    processed: list[Path] = []
    for json_path in json_files:
        try:
            payload = load_json(json_path)
        except Exception as exc:
            print(f"Skipping {json_path.name}: {exc}")
            continue

        row_data = build_row(headers, payload)
        append_row(sheet, headers, row_data)
        processed.append(json_path)
        print(f"Appended {json_path.name}")

    if processed:
        wb.save(WORKBOOK_PATH)
        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        for path in processed:
            destination = PROCESSED_DIR / path.name
            try:
                shutil.move(str(path), destination)
            except Exception as exc:
                print(f"Warning: failed to move {path.name}: {exc}")
        print(f"Appended {len(processed)} listing(s) to {WORKBOOK_PATH}")
    else:
        print("No listings appended.")


if __name__ == "__main__":
    process_queue()
