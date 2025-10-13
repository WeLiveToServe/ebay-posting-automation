"""
Utility to append agent JSON output into the eBay Seller Hub workbook.

Usage:
    python json_to_ebay_excel.py --json outputs-JSON/response_*.json
                                 --excel ebay-midsummer-upload-4.xlsx
                                 --output ebay-midsummer-upload-4.xlsx
                                 --start-price 24.99
                                 --image-url https://...

By default the script copies `--excel` to `--output` (unless `--in-place`
is supplied) and appends a new row on the `Listings` sheet using values
drawn from the JSON plus configured defaults.
"""

from __future__ import annotations

import argparse
import json
import shutil
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

import warnings

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)

from openpyxl.cell.text import Font as CommentFont  # type: ignore
from openpyxl.styles.fonts import Font as StyleFont  # type: ignore
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Expand the permitted font family range so the eBay template will load.
StyleFont.__dict__["family"].max = 100  # type: ignore[attr-defined]
CommentFont.__dict__["family"].max = 100  # type: ignore[attr-defined]

import openpyxl  # noqa: E402  (must be imported after descriptor patches)


DEFAULTS = {
    "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)": "Add",
    "Category ID": "29223",
    "Category name": "/Books & Magazines/Books",
    "Start price": "",
    "Quantity": "1",
    "Condition ID": "3000",
    "Format": "FixedPrice",
    "Duration": "GTC",
    "Location": "Los Angeles, CA",
    "Max dispatch time": "3",
    "Returns accepted option": "ReturnsAccepted",
    "Returns within option": "Days_30",
    "Refund option": "MoneyBack",
    "Return shipping cost paid by": "Buyer",
    "Shipping profile name": "",
    "Return profile name": "",
    "Payment profile name": "",
    "C:Language": "English",
}

JSON_TO_COLUMNS = {
    "title": "Title",
    "author": "C:Author",
    "edition": "C:Edition",
    "year": "C:Publication Year",
    "publisher": "C:Publisher",
    "blurb": None,  # folded into description
    "condition": None,  # folded into description
    "details": None,  # folded into description
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--json",
        required=True,
        help="Path to agent JSON output (file or directory).",
    )
    parser.add_argument(
        "--excel",
        default="ebay-midsummer-upload-4.xlsx",
        help="Path to the source workbook template.",
    )
    parser.add_argument(
        "--output",
        help="Destination workbook path. Defaults to the source path unless --in-place is set.",
    )
    parser.add_argument(
        "--sheet",
        default="Listings",
        help="Worksheet name that receives listing rows.",
    )
    parser.add_argument(
        "--start-price",
        type=str,
        default=None,
        help="Start price for the listing (leave empty to fill later).",
    )
    parser.add_argument(
        "--quantity",
        type=int,
        default=None,
        help="Quantity available (defaults to 1).",
    )
    parser.add_argument(
        "--condition-id",
        type=str,
        default=None,
        help="eBay numeric condition ID (e.g. 3000 for Used, 5000 for Good).",
    )
    parser.add_argument(
        "--category-id",
        type=str,
        default=None,
        help="eBay leaf category ID for the listing.",
    )
    parser.add_argument(
        "--category-name",
        type=str,
        default=None,
        help="Friendly category path (optional).",
    )
    parser.add_argument(
        "--image-url",
        type=str,
        default=None,
        help="Primary image URL hosted for eBay ingestion.",
    )
    parser.add_argument(
        "--location",
        type=str,
        default=None,
        help="Item location text (city, region).",
    )
    parser.add_argument(
        "--shipping-profile",
        type=str,
        default=None,
        help="Name of the eBay business policy for shipping.",
    )
    parser.add_argument(
        "--return-profile",
        type=str,
        default=None,
        help="Name of the eBay business policy for returns.",
    )
    parser.add_argument(
        "--payment-profile",
        type=str,
        default=None,
        help="Name of the eBay business policy for payment.",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Modify the source workbook directly instead of copying first.",
    )
    return parser.parse_args()


def find_latest_json(path: Path) -> Path:
    if path.is_file():
        return path
    candidates = sorted(path.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"No JSON files found in {path}")
    return candidates[0]


def load_json_payload(json_path: Path) -> Mapping[str, Any]:
    with json_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, Mapping):
        raise ValueError("JSON payload must be an object with key/value pairs.")
    return payload


def normalize_text(text: str | None) -> str:
    if not text:
        return ""
    replacements = {
        "\uFFFD": "'",
        "’": "'",
        "“": '"',
        "”": '"',
        "–": "-",
        "—": "-",
        "…": "...",
        "•": "-",
        "©": "(c)",
        "Ã©": "é",
        "Ã¼": "ü",
        "Ã¶": "ö",
        "Ã": "à",
        "�": "'",
    }
    cleaned = text
    for bad, good in replacements.items():
        cleaned = cleaned.replace(bad, good)
    return cleaned.strip()


def build_description(payload_lower: Mapping[str, Any]) -> str:
    blurb = normalize_text(payload_lower.get("blurb"))
    condition = normalize_text(payload_lower.get("condition"))
    details = normalize_text(payload_lower.get("details"))

    sections: list[str] = []
    if blurb:
        sections.append(blurb)
    if condition:
        sections.append(f"Condition Notes:\n{condition}")
    if details:
        sections.append(f"Collector Details:\n{details}")

    return "\n\n".join(sections)


def read_header_row(sheet: Worksheet) -> tuple[int, list[str]]:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        first_value = row[0].value
        if isinstance(first_value, str) and first_value.startswith("*Action("):
            headers = [cell.value if isinstance(cell.value, str) else "" for cell in row]
            return row[0].row, headers
    raise ValueError("Could not locate the header row containing '*Action'.")


def find_next_row(sheet: Worksheet, title_col_idx: int, start_row: int) -> int:
    row_index = start_row + 1
    while True:
        cell = sheet.cell(row=row_index, column=title_col_idx + 1)
        if cell.value in (None, ""):
            return row_index
        row_index += 1


def prepare_row_values(
    headers: Iterable[str],
    payload: Mapping[str, Any],
    payload_lower: Mapping[str, Any],
    args: argparse.Namespace,
    baseline: Mapping[str, Any] | None,
) -> dict[str, Any]:
    values: dict[str, Any] = {header: "" for header in headers}

    # Apply defaults first.
    for key, default_value in DEFAULTS.items():
        if key in values:
            values[key] = default_value

    # Adopt baseline for optional policy columns if present.
    if baseline:
        for policy_key in ("Shipping profile name", "Return profile name", "Payment profile name"):
            if policy_key in values and baseline.get(policy_key):
                values[policy_key] = baseline[policy_key]

    # JSON field mapping.
    for json_key, column_name in JSON_TO_COLUMNS.items():
        if column_name and column_name in values:
            values[column_name] = normalize_text(payload_lower.get(json_key))

    # Mirror title into C:Book Title when present.
    if "Title" in values and "C:Book Title" in values:
        values["C:Book Title"] = values["Title"]

    # Description block.
    if "*Description" in values:
        values["*Description"] = build_description(payload_lower)
    elif "Description" in values:
        values["Description"] = build_description(payload_lower)

    # User overrides from CLI.
    if args.start_price is not None and "Start price" in values:
        values["Start price"] = args.start_price

    if args.quantity is not None and "Quantity" in values:
        values["Quantity"] = str(args.quantity)

    if args.condition_id is not None:
        column = "Condition ID" if "Condition ID" in values else "*ConditionID"
        if column in values:
            values[column] = args.condition_id

    if args.category_id is not None:
        column = "Category ID" if "Category ID" in values else "*Category"
        if column in values:
            values[column] = args.category_id

    if args.category_name and "Category name" in values:
        values["Category name"] = args.category_name

    if args.image_url and "Item photo URL" in values:
        values["Item photo URL"] = args.image_url
    elif args.image_url and "PicURL" in values:
        values["PicURL"] = args.image_url

    if args.location and "Location" in values:
        values["Location"] = args.location

    if args.shipping_profile and "Shipping profile name" in values:
        values["Shipping profile name"] = args.shipping_profile
    if args.return_profile and "Return profile name" in values:
        values["Return profile name"] = args.return_profile
    if args.payment_profile and "Payment profile name" in values:
        values["Payment profile name"] = args.payment_profile

    # Condition text fallback.
    if payload.get("condition") and "Condition ID" not in values and "*ConditionID" in values:
        values["*ConditionID"] = args.condition_id or DEFAULTS.get("*ConditionID", "")

    return values


def derive_baseline(sheet: Worksheet, header_row: int, headers: list[str]) -> Mapping[str, Any] | None:
    title_idx = headers.index("Title") if "Title" in headers else None
    if title_idx is None:
        return None
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        title_value = row[title_idx] if title_idx < len(row) else None
        if title_value not in (None, ""):
            return {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
    return None


def apply_row(sheet: Worksheet, row_index: int, headers: list[str], values: Mapping[str, Any]) -> None:
    for col_index, header in enumerate(headers, start=1):
        if not header:
            continue
        value = values.get(header, "")
        cell = sheet.cell(row=row_index, column=col_index)
        if isinstance(value, str) and value == "":
            cell.value = None
            continue

        if header.lower().endswith("price"):
            try:
                cell.value = float(value) if value != "" else None
                continue
            except (TypeError, ValueError):
                pass

        if header.lower() == "quantity":
            try:
                cell.value = int(value)
                continue
            except (TypeError, ValueError):
                pass

        cell.value = value


def main() -> None:
    args = parse_args()

    json_path = find_latest_json(Path(args.json))
    payload = load_json_payload(json_path)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_path}")

    if args.in_place:
        output_path = excel_path
    else:
        output_path = Path(args.output) if args.output else excel_path.with_name(
            excel_path.stem + "-updated" + excel_path.suffix
        )
        shutil.copy2(excel_path, output_path)

    wb: Workbook = openpyxl.load_workbook(output_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Worksheet '{args.sheet}' not found in workbook.")

    sheet = wb[args.sheet]
    header_row_index, headers = read_header_row(sheet)
    baseline = derive_baseline(sheet, header_row_index, headers)

    title_col_idx = headers.index("Title") if "Title" in headers else 4
    target_row = find_next_row(sheet, title_col_idx, header_row_index)
    payload_lower = {str(key).lower(): value for key, value in payload.items()}

    row_values = prepare_row_values(headers, payload, payload_lower, args, baseline)
    apply_row(sheet, target_row, headers, row_values)

    wb.save(output_path)
    print(f"Appended listing to '{output_path}' on row {target_row}.")


if __name__ == "__main__":
    main()
